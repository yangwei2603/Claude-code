from datetime import date
from typing import Optional

from fastapi import APIRouter, Depends, Query
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from io import BytesIO

from app.database import get_db
from app.models import Requirement, RequirementReview
from app.services.workdays import parse_date

router = APIRouter()

EXPORT_FIELDS = [
    "demand_id",
    "demand_name",
    "priority",
    "decision",
    "review_date",
    "description",
    "reason",
]


def _decision(rv: RequirementReview) -> str:
    if rv.plan_status == "未来规划":
        return "未来规划"
    if rv.rejection_type:
        return "驳回"
    return "通过"


def _build_rows(
    db: Session,
    date_from: Optional[str],
    date_to: Optional[str],
    decision_filter: Optional[str],
):
    q = db.query(RequirementReview, Requirement).join(
        Requirement, RequirementReview.requirement_id == Requirement.demand_id
    )
    if date_from:
        q = q.filter(RequirementReview.review_date >= date_from)
    if date_to:
        q = q.filter(RequirementReview.review_date <= date_to)

    rows = []
    for rv, req in q.all():
        dec = _decision(rv)
        if decision_filter and dec != decision_filter:
            continue
        rows.append(
            {
                "demand_id": req.demand_id,
                "demand_name": req.demand_name,
                "priority": rv.priority or "",
                "decision": dec,
                "review_date": rv.review_date or "",
                "description": rv.review_notes or "",
                "reason": rv.rejection_reason or "",
            }
        )
    return rows


@router.get("/reviews/export-preview")
def export_preview(
    date_from: Optional[str] = None,
    date_to: Optional[str] = None,
    decision: Optional[str] = None,
    db: Session = Depends(get_db),
):
    rows = _build_rows(db, date_from, date_to, decision)
    return {"code": 0, "message": "success", "data": rows}


@router.get("/reviews/export")
def export_excel(
    date_from: Optional[str] = None,
    date_to: Optional[str] = None,
    decision: Optional[str] = None,
    db: Session = Depends(get_db),
):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    rows = _build_rows(db, date_from, date_to, decision)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "评审导出"

    headers = ["编号", "名称", "优先级", "评审分类", "上线时间", "描述", "原因"]
    header_fill = PatternFill("solid", fgColor="4472C4")
    header_font = Font(color="FFFFFF", bold=True)
    thin = Side(style="thin")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
        cell.border = border

    for row_idx, row in enumerate(rows, 2):
        ws.cell(row=row_idx, column=1, value=row["demand_id"]).border = border
        ws.cell(row=row_idx, column=2, value=row["demand_name"]).border = border
        ws.cell(row=row_idx, column=3, value=row["priority"]).border = border
        ws.cell(row=row_idx, column=4, value=row["decision"]).border = border
        ws.cell(row=row_idx, column=5, value=row["review_date"]).border = border
        ws.cell(row=row_idx, column=6, value=row["description"]).border = border
        ws.cell(row=row_idx, column=7, value=row["reason"]).border = border

    for col in range(1, 8):
        ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = 20

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)

    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": "attachment; filename=review_export.xlsx"
        },
    )