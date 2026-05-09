"""数据导出模块"""
import csv
from datetime import date
from pathlib import Path
from typing import List, Dict, Any

from .models import MarketData, MemberPosition, TradingSignal


class CSVExporter:
    """CSV导出器"""

    def export_market_data(self, data_list: List[MarketData], filepath: str) -> None:
        """导出行情数据到CSV"""
        if not data_list:
            return

        fieldnames = [
            "contract_code", "trade_date", "close_price",
            "volume", "open_interest",
            "prev_close_price", "prev_volume", "prev_open_interest"
        ]

        with open(filepath, "w", newline="", encoding="utf-8-sig") as f:
            writer = csv.DictWriter(f, fieldnames=fieldnames)
            writer.writeheader()
            for data in data_list:
                writer.writerow({
                    "contract_code": data.contract_code,
                    "trade_date": data.trade_date.isoformat(),
                    "close_price": round(data.close_price, 2),
                    "volume": data.volume,
                    "open_interest": data.open_interest,
                    "prev_close_price": round(data.prev_close_price, 2),
                    "prev_volume": data.prev_volume,
                    "prev_open_interest": data.prev_open_interest,
                })

    def export_positions(self, data_list: List[MemberPosition], filepath: str) -> None:
        """导出持仓数据到CSV"""
        if not data_list:
            return

        fieldnames = [
            "seat_code", "seat_name", "contract_code", "trade_date",
            "long_position", "short_position",
            "prev_long_position", "prev_short_position"
        ]

        with open(filepath, "w", newline="", encoding="utf-8-sig") as f:
            writer = csv.DictWriter(f, fieldnames=fieldnames)
            writer.writeheader()
            for data in data_list:
                writer.writerow({
                    "seat_code": data.seat_code,
                    "seat_name": data.seat_name,
                    "contract_code": data.contract_code,
                    "trade_date": data.trade_date.isoformat(),
                    "long_position": data.long_position,
                    "short_position": data.short_position,
                    "prev_long_position": data.prev_long_position,
                    "prev_short_position": data.prev_short_position,
                })

    def export_signals(self, signals: List[TradingSignal], filepath: str) -> None:
        """导出信号到CSV"""
        if not signals:
            return

        fieldnames = [
            "trade_date", "signal_type", "signal_name",
            "contract_code", "seat_code", "confidence"
        ]

        with open(filepath, "w", newline="", encoding="utf-8-sig") as f:
            writer = csv.DictWriter(f, fieldnames=fieldnames)
            writer.writeheader()
            for sig in signals:
                writer.writerow({
                    "trade_date": sig.trade_date.isoformat(),
                    "signal_type": sig.signal_type,
                    "signal_name": sig.signal_name,
                    "contract_code": sig.contract_code,
                    "seat_code": sig.seat_code,
                    "confidence": sig.confidence,
                })


class ExcelExporter:
    """Excel导出器"""

    def __init__(self):
        try:
            import openpyxl
            self.openpyxl = openpyxl
        except ImportError:
            self.openpyxl = None

    def _ensure_openpyxl(self):
        if self.openpyxl is None:
            raise ImportError("openpyxl is required for Excel export. Install with: pip install openpyxl")

    def export_market_data(self, data_list: List[MarketData], filepath: str) -> None:
        """导出行情数据到Excel"""
        self._ensure_openpyxl()
        if not data_list:
            return

        wb = self.openpyxl.Workbook()
        ws = wb.active
        ws.title = "行情数据"

        headers = [
            "合约代码", "交易日期", "收盘价",
            "成交量", "持仓量", "前收盘价", "前成交量", "前持仓量"
        ]
        ws.append(headers)

        for data in data_list:
            ws.append([
                data.contract_code,
                data.trade_date.isoformat(),
                round(data.close_price, 2),
                data.volume,
                data.open_interest,
                round(data.prev_close_price, 2),
                data.prev_volume,
                data.prev_open_interest,
            ])

        wb.save(filepath)

    def export_positions(self, data_list: List[MemberPosition], filepath: str) -> None:
        """导出持仓数据到Excel"""
        self._ensure_openpyxl()
        if not data_list:
            return

        wb = self.openpyxl.Workbook()
        ws = wb.active
        ws.title = "会员持仓"

        headers = [
            "席位代码", "席位名称", "合约代码", "交易日期",
            "多头持仓", "空头持仓", "前多头持仓", "前空头持仓"
        ]
        ws.append(headers)

        for data in data_list:
            ws.append([
                data.seat_code,
                data.seat_name,
                data.contract_code,
                data.trade_date.isoformat(),
                data.long_position,
                data.short_position,
                data.prev_long_position,
                data.prev_short_position,
            ])

        wb.save(filepath)

    def export_signals(self, signals: List[TradingSignal], filepath: str) -> None:
        """导出信号到Excel"""
        self._ensure_openpyxl()
        if not signals:
            return

        wb = self.openpyxl.Workbook()
        ws = wb.active
        ws.title = "交易信号"

        headers = ["交易日期", "信号类型", "信号名称", "合约代码", "席位代码", "置信度"]
        ws.append(headers)

        for sig in signals:
            ws.append([
                sig.trade_date.isoformat(),
                sig.signal_type,
                sig.signal_name,
                sig.contract_code,
                sig.seat_code,
                sig.confidence,
            ])

        wb.save(filepath)

    def export_all(
        self,
        market_data: Dict[str, List[MarketData]],
        positions: Dict[str, List[MemberPosition]],
        signals: List[TradingSignal],
        filepath: str
    ) -> None:
        """导出所有数据到单个Excel文件"""
        self._ensure_openpyxl()

        wb = self.openpyxl.Workbook()

        # 行情数据表
        ws_market = wb.create_sheet("行情数据")
        ws_market.append(["合约代码", "交易日期", "收盘价", "成交量", "持仓量"])
        for contract, data_list in market_data.items():
            for data in data_list:
                ws_market.append([
                    data.contract_code,
                    data.trade_date.isoformat(),
                    round(data.close_price, 2),
                    data.volume,
                    data.open_interest,
                ])

        # 持仓数据表
        ws_position = wb.create_sheet("会员持仓")
        ws_position.append(["席位代码", "席位名称", "合约代码", "交易日期", "多头持仓", "空头持仓"])
        for key, data_list in positions.items():
            for data in data_list:
                ws_position.append([
                    data.seat_code,
                    data.seat_name,
                    data.contract_code,
                    data.trade_date.isoformat(),
                    data.long_position,
                    data.short_position,
                ])

        # 信号表
        ws_signals = wb.create_sheet("交易信号")
        ws_signals.append(["交易日期", "信号类型", "信号名称", "合约代码", "席位代码"])
        for sig in signals:
            ws_signals.append([
                sig.trade_date.isoformat(),
                sig.signal_type,
                sig.signal_name,
                sig.contract_code,
                sig.seat_code,
            ])

        # 删除默认的Sheet
        if "Sheet" in wb.sheetnames:
            del wb["Sheet"]

        wb.save(filepath)


def get_exporter(format: str):
    """获取指定的导出器"""
    if format.lower() == "csv":
        return CSVExporter()
    elif format.lower() in ("excel", "xlsx"):
        return ExcelExporter()
    else:
        raise ValueError(f"Unsupported format: {format}. Use 'csv' or 'xlsx'")
